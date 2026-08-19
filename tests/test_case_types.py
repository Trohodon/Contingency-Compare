import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from core.case_finder import _classify_case
from core.case_types import CASE_TYPE_DEFINITIONS
from core.comparator import (
    build_all_case_type_comparisons,
    build_batch_comparison_workbook,
)
from core.comparison_builder import build_workbook
from core.column_blacklist import apply_blacklist, apply_row_filter, VOLTAGE_VIOLATION_CATEGORIES
from core.case_processor import post_process_csv


def _write_scenario_sheet(ws, offset):
    row = 2
    for canonical, pretty in CASE_TYPE_DEFINITIONS:
        ws.cell(row=row, column=2).value = pretty
        row += 1
        headers = [
            "Contingency Events",
            "Resulting Issue",
            "Limit",
            "Contingency Value (MVA)",
            "Percent Loading",
        ]
        for col, header in enumerate(headers, start=2):
            ws.cell(row=row, column=col).value = header
        row += 1

        ws.cell(row=row, column=2).value = f"{canonical} CTG"
        ws.cell(row=row, column=3).value = f"{canonical} Issue"
        ws.cell(row=row, column=4).value = 100
        ws.cell(row=row, column=5).value = 90 + offset
        ws.cell(row=row, column=6).value = 90 + offset
        row += 1

        if canonical == "AUXapplied":
            ws.cell(row=row, column=2).value = f"{canonical} CTG"
            ws.cell(row=row, column=3).value = f"{canonical} Issue"
            ws.cell(row=row, column=4).value = 100
            ws.cell(row=row, column=5).value = 95 + offset
            ws.cell(row=row, column=6).value = f"{95 + offset}%"
            row += 1

        row += 1


class CaseTypeIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.source_path = os.path.join(self.temp_dir.name, "source.xlsx")

        wb = Workbook()
        left = wb.active
        left.title = "Left"
        right = wb.create_sheet("Right")
        _write_scenario_sheet(left, 0)
        _write_scenario_sheet(right, 5)
        wb.save(self.source_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_auxapplied_filename_recognition_is_case_insensitive(self):
        self.assertEqual(_classify_case("Study_AUXapplied_Final.pwb"), "AUXapplied")
        self.assertEqual(_classify_case("study_auxapplied_final.PWB"), "AUXapplied")

    def test_all_case_types_compare_and_duplicate_keys_use_maximum(self):
        comparisons = build_all_case_type_comparisons(
            self.source_path, "Left", "Right"
        )

        self.assertEqual(set(comparisons), {c for c, _p in CASE_TYPE_DEFINITIONS})
        aux = comparisons["AUXapplied"]
        self.assertEqual(len(aux), 1)
        self.assertEqual(aux.iloc[0]["LeftPct"], 95)
        self.assertEqual(aux.iloc[0]["RightPct"], 100)
        self.assertEqual(aux.iloc[0]["DeltaPct"], 5)

    def test_batch_and_straight_outputs_include_auxapplied(self):
        output_path = os.path.join(self.temp_dir.name, "comparison.xlsx")
        build_batch_comparison_workbook(
            src_workbook=self.source_path,
            pairs=[("Left", "Right")],
            threshold=80,
            output_path=output_path,
        )

        wb = load_workbook(output_path, read_only=True, data_only=True)
        self.assertIn("Left vs Right", wb.sheetnames)
        self.assertIn("Straight Comparison", wb.sheetnames)

        pair_titles = [
            row[0]
            for row in wb["Left vs Right"].iter_rows(
                min_col=2, max_col=2, values_only=True
            )
            if row[0]
        ]
        straight_titles = [
            row[0]
            for row in wb["Straight Comparison"].iter_rows(
                min_col=2, max_col=2, values_only=True
            )
            if row[0]
        ]
        self.assertIn("AUXapplied", pair_titles)
        self.assertIn("AUXapplied", straight_titles)
        wb.close()

    def test_high_threshold_still_creates_valid_workbook(self):
        output_path = os.path.join(self.temp_dir.name, "empty.xlsx")
        build_batch_comparison_workbook(
            src_workbook=self.source_path,
            pairs=[("Left", "Right")],
            threshold=500,
            output_path=output_path,
        )

        wb = load_workbook(output_path, read_only=True, data_only=True)
        self.assertEqual(
            wb["Left vs Right"].cell(row=4, column=2).value,
            "None",
        )
        self.assertEqual(
            wb["Left vs Right"].cell(row=4, column=3).value,
            "No Voltage Issues",
        )
        wb.close()

    def test_blacklist_keeps_limit_column_needed_by_workbooks(self):
        import pandas as pd

        df = pd.DataFrame(
            {
                "CTGLabel": ["CTG"],
                "LimViolID": ["Issue"],
                "LimViolLimit": [100],
                "LimViolValue": [90],
                "LimViolPct": [90],
                "LimViolCat": ["Branch MVA"],
                "BusNum": [1],
            }
        )
        out, removed = apply_blacklist(df)
        self.assertIn("LimViolLimit", out.columns)
        self.assertIn("LimViolCat", out.columns)
        self.assertNotIn("BusNum", out.columns)
        self.assertIn("BusNum", removed)

    def test_combined_workbook_threshold_writes_no_voltage_issues(self):
        import pandas as pd

        csv_path = os.path.join(self.temp_dir.name, "case.csv")
        pd.DataFrame(
            {
                "CTGLabel": ["CTG"],
                "LimViolID": ["Issue"],
                "LimViolLimit": [100],
                "LimViolValue": [50],
                "LimViolPct": [50],
            }
        ).to_csv(csv_path, index=False)

        output_path = build_workbook(
            self.temp_dir.name,
            {"Scenario": {"ACCA_P1,2,4,7": csv_path}},
            threshold=80,
        )

        wb = load_workbook(output_path, read_only=True, data_only=True)
        ws = wb["Scenario"]
        self.assertEqual(ws.cell(row=4, column=2).value, "None")
        self.assertEqual(ws.cell(row=4, column=3).value, "No Voltage Issues")
        self.assertEqual(ws.cell(row=3, column=7).value, "Notes")
        wb.close()

    def test_no_violation_export_still_creates_filtered_csv(self):
        import pandas as pd

        csv_path = os.path.join(self.temp_dir.name, "empty_violation_ctg.csv")
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            f.write("ViolationCTG\n")
            f.write("CTGLabel,LimViolID,LimViolLimit,LimViolValue,LimViolPct,LimViolCat\n")

        filtered_path = post_process_csv(
            csv_path,
            dedup_enabled=True,
            keep_categories={"Branch MVA"},
            threshold=100,
        )

        self.assertTrue(os.path.isfile(filtered_path))
        filtered = pd.read_csv(filtered_path)
        self.assertEqual(len(filtered), 0)
        self.assertIn("LimViolPct", filtered.columns)

    def test_title_only_no_violation_export_still_creates_filtered_csv(self):
        import pandas as pd

        csv_path = os.path.join(self.temp_dir.name, "title_only_violation_ctg.csv")
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            f.write("ViolationCTG\n")

        filtered_path = post_process_csv(
            csv_path,
            dedup_enabled=True,
            keep_categories={"Branch MVA"},
            threshold=100,
        )

        self.assertTrue(os.path.isfile(filtered_path))
        filtered = pd.read_csv(filtered_path)
        self.assertEqual(len(filtered), 0)
        self.assertIn("LimViolPct", filtered.columns)

    def test_empty_no_violation_export_still_creates_filtered_csv(self):
        import pandas as pd

        csv_path = os.path.join(self.temp_dir.name, "empty_file_violation_ctg.csv")
        open(csv_path, "w", encoding="utf-8").close()

        filtered_path = post_process_csv(
            csv_path,
            dedup_enabled=True,
            keep_categories={"Branch MVA"},
            threshold=100,
        )

        self.assertTrue(os.path.isfile(filtered_path))
        filtered = pd.read_csv(filtered_path)
        self.assertEqual(len(filtered), 0)
        self.assertIn("LimViolPct", filtered.columns)

    def test_voltage_filter_includes_low_high_and_change_high_voltage(self):
        import pandas as pd

        df = pd.DataFrame(
            {
                "LimViolCat": [
                    "Bus Low Volts",
                    "Bus High Volts",
                    "Change Bus High Volts",
                    "Branch MVA",
                    "Other",
                ],
                "LimViolPct": [101, 102, 103, 104, 105],
            }
        )

        filtered, removed = apply_row_filter(df, keep_values=VOLTAGE_VIOLATION_CATEGORIES)

        self.assertEqual(removed, 2)
        self.assertEqual(
            set(filtered["LimViolCat"]),
            {"Bus Low Volts", "Bus High Volts", "Change Bus High Volts"},
        )

    def test_voltage_workbook_ignores_percent_threshold_and_uses_pu_hundredths(self):
        import pandas as pd

        csv_path = os.path.join(self.temp_dir.name, "voltage_case.csv")
        pd.DataFrame(
            {
                "CTGLabel": ["Voltage CTG"],
                "LimViolID": ["Voltage Issue"],
                "LimViolLimit": [0.956],
                "LimViolValue": [1.047],
                "LimViolPct": [50],
                "LimViolCat": ["Bus High Volts"],
            }
        ).to_csv(csv_path, index=False)

        output_path = build_workbook(
            self.temp_dir.name,
            {"Scenario": {"ACCA_LongTerm": csv_path}},
            threshold=100,
            report_type="voltage",
        )

        wb = load_workbook(output_path, read_only=False, data_only=True)
        try:
            ws = wb["Scenario"]
            self.assertEqual(ws.cell(row=3, column=5).value, "Contingency Value (p.u.)")
            self.assertTrue(ws.column_dimensions["F"].hidden)
            self.assertEqual(ws.cell(row=4, column=2).value, "Voltage CTG")
            self.assertEqual(ws.cell(row=4, column=4).value, 0.96)
            self.assertEqual(ws.cell(row=4, column=5).value, 1.05)
            self.assertIsNone(ws.cell(row=4, column=6).value)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
