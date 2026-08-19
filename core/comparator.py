# core/comparator.py
#
# Helpers for working with the formatted
# Combined_ViolationCTG_Comparison.xlsx workbook and for building
# batch comparison workbooks in a nicely formatted style.
#
# Public functions used by the GUI:
#   - list_sheets(workbook_path)
#   - build_case_type_comparison(... )
#   - build_pair_comparison_df(... )
#   - build_batch_comparison_workbook(... )
#
# UPDATED:
#   - Parsing supports BOTH old formatted sheets (no Limit column) and new ones (with Limit).
#   - Pair + Straight Comparison outputs can include Limit.
#   - build_batch_comparison_workbook allows pairs=[] so users can build a workbook
#     containing ONLY the "Straight Comparison" sheet (all originals).

from __future__ import annotations

from typing import List, Dict, Optional, Sequence, Tuple

import math
import os
import pandas as pd

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.batch_sheet_writer import write_formatted_pair_sheet
from core.case_types import (
CANONICAL_TO_PRETTY,
    CASE_TYPES_CANONICAL,
    PRETTY_TO_CANONICAL,
)
from core.straight_comparison import build_straight_comparison_df, write_formatted_straight_sheet


PARSED_COLUMNS = [
    "CaseType",
    "CTGLabel",
    "LimViolID",
    "LimViolLimit",
    "LimViolValue",
    "LimViolPct",
]


def list_sheets(workbook_path: str) -> List[str]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for sheet listing and comparison.")
    if not os.path.isfile(workbook_path):
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _header_has_limit(ws, header_row: int) -> bool:
    """Return True if the formatted header row includes a 'Limit' column in D."""
    try:
        d = ws.cell(row=header_row, column=4).value
        if isinstance(d, str) and "limit" in d.strip().lower():
            return True
    except Exception:
        pass
    return False


def _parse_scenario_sheet(ws, log_func=None) -> pd.DataFrame:
    """
    Parse one formatted scenario sheet into rows with:
        CaseType, CTGLabel, LimViolID, LimViolLimit, LimViolValue, LimViolPct

    Supports:
      - OLD format:  B=CTGLabel, C=LimViolID, D=LimViolValue, E=LimViolPct
      - NEW format:  B=CTGLabel, C=LimViolID, D=LimViolLimit, E=LimViolValue, F=LimViolPct
    """
    records: List[Dict] = []

    max_row = ws.max_row or 1
    row_idx = 1

    while row_idx <= max_row:
        title_cell = ws.cell(row=row_idx, column=2)
        title_val = title_cell.value

        if isinstance(title_val, str) and title_val.strip():
            pretty_name = title_val.strip()
            case_type = PRETTY_TO_CANONICAL.get(pretty_name, pretty_name)

            header_row = row_idx + 1
            data_row = header_row + 1

            has_limit = _header_has_limit(ws, header_row)

            last_issue = None
            r = data_row

            while r <= max_row:
                b = ws.cell(row=r, column=2).value
                c = ws.cell(row=r, column=3).value

                if has_limit:
                    lim = ws.cell(row=r, column=4).value
                    val = ws.cell(row=r, column=5).value
                    pct = ws.cell(row=r, column=6).value
                    blank_line = _is_blank(b) and _is_blank(c) and _is_blank(lim) and _is_blank(val) and _is_blank(pct)
                else:
                    lim = None
                    val = ws.cell(row=r, column=4).value
                    pct = ws.cell(row=r, column=5).value
                    blank_line = _is_blank(b) and _is_blank(c) and _is_blank(val) and _is_blank(pct)

                if blank_line:
                    break

                # Forward fill issue if blanks are used for visual grouping
                if _is_blank(c) and last_issue is not None:
                    c = last_issue
                else:
                    if not _is_blank(c):
                        last_issue = c

                records.append(
                    {
                        "CaseType": case_type,
                        "CTGLabel": b,
                        "LimViolID": c,
                        "LimViolLimit": lim,
                        "LimViolValue": val,
                        "LimViolPct": pct,
                    }
                )
                r += 1

            row_idx = r + 1  # move past block + blank separator row
        else:
            row_idx += 1

    df = pd.DataFrame.from_records(records, columns=PARSED_COLUMNS)
    if log_func:
        log_func(
            f"Parsed {len(df)} rows from sheet '{ws.title}'. "
            f"Columns: {list(df.columns)}"
        )
    return df


def _load_sheet_as_df(workbook_path: str, sheet_name: str, log_func=None) -> pd.DataFrame:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for comparison.")
    wb = load_workbook(workbook_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        return _parse_scenario_sheet(wb[sheet_name], log_func=log_func)
    finally:
        wb.close()


def _load_sheet_pair(
    workbook_path: str,
    base_sheet: str,
    new_sheet: str,
    log_func=None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for comparison.")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        missing = [name for name in (base_sheet, new_sheet) if name not in wb.sheetnames]
        if missing:
            raise ValueError(f"Sheet(s) not found: {', '.join(missing)}")
        base_df = _parse_scenario_sheet(wb[base_sheet], log_func=log_func)
        if new_sheet == base_sheet:
            new_df = base_df.copy()
        else:
            new_df = _parse_scenario_sheet(wb[new_sheet], log_func=log_func)
        return base_df, new_df
    finally:
        wb.close()


def _build_case_type_comparison_from_frames(
    base_df: pd.DataFrame,
    new_df: pd.DataFrame,
    case_type: str,
    max_rows: Optional[int] = None,
    log_func=None,
) -> pd.DataFrame:
    if case_type not in CASE_TYPES_CANONICAL:
        raise ValueError(f"Unknown case type: {case_type}")

    base_df = base_df[base_df["CaseType"] == case_type].copy()
    new_df = new_df[new_df["CaseType"] == case_type].copy()

    if log_func:
        log_func(f"  [{case_type}] base rows={len(base_df)}, new rows={len(new_df)}")

    if base_df.empty and new_df.empty:
        return pd.DataFrame(
            columns=["Contingency", "ResultingIssue", "Limit", "LeftPct", "RightPct", "DeltaPct"]
        )

    def prepare_side(df: pd.DataFrame, pct_name: str, limit_name: str) -> pd.DataFrame:
        side = df[["CTGLabel", "LimViolID", "LimViolPct", "LimViolLimit"]].copy()
        side["LimViolPct"] = pd.to_numeric(
            side["LimViolPct"].astype(str).str.replace("%", "", regex=False).str.strip(),
            errors="coerce",
        )
        side["_row_order"] = range(len(side))
        side = side.sort_values(
            by=["LimViolPct", "_row_order"],
            ascending=[False, True],
            na_position="last",
            kind="mergesort",
        )
        side = side.drop_duplicates(subset=["CTGLabel", "LimViolID"], keep="first")
        return side.rename(
            columns={"LimViolPct": pct_name, "LimViolLimit": limit_name}
        ).drop(columns=["_row_order"])

    base_df = prepare_side(base_df, "Left_Pct", "Left_Limit")
    new_df = prepare_side(new_df, "Right_Pct", "Right_Limit")

    key_cols = ["CTGLabel", "LimViolID"]
    left_cols = key_cols + ["Left_Pct", "Left_Limit"]
    right_cols = key_cols + ["Right_Pct", "Right_Limit"]

    merged = pd.merge(base_df[left_cols], new_df[right_cols], on=key_cols, how="outer")

    # Pick a single Limit value (prefer Left, else Right)
    merged["Limit"] = merged.get("Left_Limit").combine_first(merged.get("Right_Limit"))

    merged["Delta_Pct"] = merged["Right_Pct"] - merged["Left_Pct"]

    result = merged.rename(
        columns={
            "CTGLabel": "Contingency",
            "LimViolID": "ResultingIssue",
            "Left_Pct": "LeftPct",
            "Right_Pct": "RightPct",
            "Delta_Pct": "DeltaPct",
        }
    )

    # Sort by whichever side has values
    sort_series = result["RightPct"]
    result["_SortPct"] = sort_series if sort_series.notna().any() else result["LeftPct"]
    result = result.sort_values(by="_SortPct", ascending=False, na_position="last").drop(columns=["_SortPct"])

    if max_rows is not None and max_rows > 0:
        result = result.head(max_rows)

    # Keep a clean column order
    keep = ["Contingency", "ResultingIssue", "Limit", "LeftPct", "RightPct", "DeltaPct"]
    for col in keep:
        if col not in result.columns:
            result[col] = None
    return result[keep].copy()


def build_case_type_comparison(
    workbook_path: str,
    base_sheet: str,
    new_sheet: str,
    case_type: str,
    max_rows: Optional[int] = None,
    log_func=None,
) -> pd.DataFrame:
    """
    Returns a DataFrame containing:
      Contingency, ResultingIssue, Limit, LeftPct, RightPct, DeltaPct
    """
    base_df, new_df = _load_sheet_pair(
        workbook_path, base_sheet, new_sheet, log_func=log_func
    )
    return _build_case_type_comparison_from_frames(
        base_df, new_df, case_type, max_rows=max_rows, log_func=log_func
    )


def build_all_case_type_comparisons(
    workbook_path: str,
    base_sheet: str,
    new_sheet: str,
    max_rows: Optional[int] = None,
    log_func=None,
) -> Dict[str, pd.DataFrame]:
    """Parse two sheets once and return a comparison for every case type."""
    base_df, new_df = _load_sheet_pair(
        workbook_path, base_sheet, new_sheet, log_func=log_func
    )
    return {
        case_type: _build_case_type_comparison_from_frames(
            base_df,
            new_df,
            case_type,
            max_rows=max_rows,
            log_func=log_func,
        )
        for case_type in CASE_TYPES_CANONICAL
    }


def _is_nan(x) -> bool:
    return isinstance(x, float) and math.isnan(x)


def build_pair_comparison_df(
    workbook_path: str,
    left_sheet: str,
    right_sheet: str,
    threshold: float,
    log_func=None,
) -> pd.DataFrame:
    records: List[Dict] = []
    comparisons = build_all_case_type_comparisons(
        workbook_path,
        left_sheet,
        right_sheet,
        max_rows=None,
        log_func=log_func,
    )

    for case_type in CASE_TYPES_CANONICAL:
        pretty = CANONICAL_TO_PRETTY.get(case_type, case_type)
        df = comparisons[case_type]
        if df.empty:
            continue

        for _, row in df.iterrows():
            cont = str(row.get("Contingency", "") or "")
            issue = "" if row.get("ResultingIssue", None) is None else str(row.get("ResultingIssue"))
            limit = row.get("Limit", None)

            left_pct = row.get("LeftPct", math.nan)
            right_pct = row.get("RightPct", math.nan)
            delta_pct = row.get("DeltaPct", math.nan)

            # Threshold filter uses the max of (left, right)
            values = []
            if not _is_nan(left_pct):
                values.append(float(left_pct))
            if not _is_nan(right_pct):
                values.append(float(right_pct))
            if not values or max(values) < float(threshold):
                continue

            if _is_nan(left_pct) and not _is_nan(right_pct):
                delta_text = "Only in right"
            elif not _is_nan(left_pct) and _is_nan(right_pct):
                delta_text = "Only in left"
            elif _is_nan(left_pct) and _is_nan(right_pct):
                delta_text = ""
            else:
                delta_text = f"{float(delta_pct):.2f}" if not _is_nan(delta_pct) else ""

            records.append(
                {
                    "CaseType": pretty,
                    "Contingency": cont,
                    "ResultingIssue": issue,
                    "Limit": limit,
                    "LeftPct": float(left_pct) if not _is_nan(left_pct) else None,
                    "RightPct": float(right_pct) if not _is_nan(right_pct) else None,
                    "DeltaDisplay": delta_text,
                }
            )

    df_all = pd.DataFrame.from_records(records)
    if not df_all.empty:
        sort_vals = df_all[["LeftPct", "RightPct"]].max(axis=1)
        df_all["_SortKey"] = sort_vals
        df_all = df_all.sort_values(
            by=["CaseType", "_SortKey"],
            ascending=[True, False],
            na_position="last",
        ).drop(columns=["_SortKey"])
    return df_all


def _sanitize_sheet_name(name: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join(ch if ch not in invalid else "_" for ch in name).strip()
    return (cleaned or "Sheet")[:31]


def _looks_like_output_sheet(name: str) -> bool:
    """Filter out sheets that are clearly generated outputs."""
    n = (name or "").strip().lower()
    if not n:
        return True
    if " vs " in n:
        return True
    if n.startswith("straight comparison"):
        return True
    if n.startswith("batch comparison"):
        return True
    if n.startswith("comparison"):
        return True
    return False


def _ordered_original_sheets(src_workbook: str, pairs: Sequence[Tuple[str, str]]) -> List[str]:
    """
    If pairs are provided: return sheets involved in pairs in workbook order.
    If pairs is empty: return ALL likely-original scenario sheets in workbook order.
    """
    try:
        wb = load_workbook(src_workbook, read_only=True, data_only=True)
        try:
            sheetnames = list(wb.sheetnames)

            # If no pairs, include all "original-looking" sheets
            if not pairs:
                originals = [s for s in sheetnames if not _looks_like_output_sheet(s)]
                return originals

            wanted: set[str] = set()
            for a, b in pairs:
                wanted.add(a)
                wanted.add(b)

            ordered = [s for s in sheetnames if s in wanted]
            for s in wanted:
                if s not in ordered:
                    ordered.append(s)
            return ordered
        finally:
            wb.close()

    except Exception:
        # Fallback behavior if workbook can't be opened
        if not pairs:
            return []

        ordered: List[str] = []
        seen: set[str] = set()
        for a, b in pairs:
            for s in (a, b):
                if s not in seen:
                    seen.add(s)
                    ordered.append(s)
        return ordered


def build_batch_comparison_workbook(
    src_workbook: Optional[str] = None,
    pairs: Sequence[Tuple[str, str]] = (),
    threshold: float = 0.0,
    output_path: str = "",
    log_func=None,
    *,
    expandable_issue_view: bool = True,
    workbook_path: Optional[str] = None,
    **kwargs,
) -> str:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to build the batch workbook.")

    if src_workbook is None:
        src_workbook = workbook_path
    if src_workbook is None:
        src_workbook = kwargs.get("src_workbook") or kwargs.get("workbook") or kwargs.get("path")
    if not src_workbook:
        raise ValueError("Missing source workbook path (src_workbook / workbook_path).")

    # pairs can be empty -> build a workbook with only Straight Comparison.
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()

    # Pair sheets (only if pairs exist)
    if pairs:
        for (left_sheet, right_sheet) in pairs:
            df_pair = build_pair_comparison_df(
                src_workbook, left_sheet, right_sheet, threshold, log_func=log_func
            )
            if df_pair.empty:
                df_pair = pd.DataFrame([{
                    "CaseType": "",
                    "Contingency": "None",
                    "ResultingIssue": "No Voltage Issues",
                    "Limit": None,
                    "LeftPct": None,
                    "RightPct": None,
                    "DeltaDisplay": "",
                    "Notes": "",
                }])

            base_name = _sanitize_sheet_name(f"{left_sheet} vs {right_sheet}")
            name = base_name
            counter = 2
            while name in used_names:
                suffix = f" ({counter})"
                name = _sanitize_sheet_name(base_name[: (31 - len(suffix))] + suffix)
                counter += 1
            used_names.add(name)

            write_formatted_pair_sheet(wb, name, df_pair, expandable_issue_view=expandable_issue_view)
    else:
        if log_func:
            log_func("No queued pairs provided; building Straight Comparison only.")

    # Straight Comparison (always attempt)
    try:
        originals = _ordered_original_sheets(src_workbook, pairs)
        if not originals:
            if log_func:
                log_func("WARNING: No original sheets detected for Straight Comparison.")
        else:
            df_straight, case_labels = build_straight_comparison_df(
                src_workbook,
                originals,
                threshold=threshold,
                log_func=log_func
            )

            sc_base = _sanitize_sheet_name("Straight Comparison")
            sc_name = sc_base
            k = 2
            while sc_name in used_names:
                suffix = f" ({k})"
                sc_name = _sanitize_sheet_name(sc_base[: (31 - len(suffix))] + suffix)
                k += 1
            used_names.add(sc_name)

            write_formatted_straight_sheet(
                wb,
                sc_name,
                df_straight,
                case_labels,
                expandable_issue_view=expandable_issue_view,
            )
    except Exception as e:
        if log_func:
            log_func(f"WARNING: Straight Comparison sheet failed: {e}")

    if not wb.sheetnames:
        ws = wb.create_sheet("Comparison")
        ws.cell(row=1, column=1).value = "No comparison data was available."

    wb.save(output_path)
    return output_path
